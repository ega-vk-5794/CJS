# -*- coding: utf-8 -*-
"""Generate SNAPSHOT.md and snapshot.html from register.py. Run: python3 build.py"""
import html, json, os, sys, datetime
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import register as R

HERE = os.path.dirname(os.path.abspath(__file__))
AS_AT = "5 September 2026"

STATE = {
 "fw":       ("Framework - we push", "fw"),
 "handler":  ("Handler class - developer", "hd"),
 "done_dev": ("Closed by developer", "done"),
 "config":   ("Config - developer", "cfg"),
 "backend":  ("Backend / record model", "be"),
 "portal":   ("Portal / launchpad", "pt"),
 "check":    ("Needs a look", "chk"),
}
ORDER = ("fw", "handler", "done_dev", "config", "backend", "portal", "check")
OWNER = {
 "fw":       "Framework team (us)",
 "handler":  "Journey developer - ABAP",
 "done_dev": "Journey developer - closed",
 "config":   "Journey developer - Studio config",
 "backend":  "Backend team",
 "portal":   "Portal team",
 "check":    "Reporter + developer",
}
CLOSED = ("fw", "done_dev")

by_j = {}
for t, seq, j, text, st, act in R.OBS:
    by_j.setdefault(j, []).append((t, seq, text, st, act))

def counts(rows):
    c = {k: 0 for k in STATE}
    for r in rows:
        c[r[3]] += 1
    return c

tot = counts([(0,0,0,o[4],0) for o in R.OBS])
n_obs = len(R.OBS)
n_closed = tot["fw"] + tot["done_dev"]
n_open = n_obs - n_closed

# ---------------------------------------------------------------- markdown
md = []
md.append(f"# CJS journey snapshot - {AS_AT}\n")
md.append("Built from the five Jira exports (Issues + Comments sheets). "
          "Regenerate with `python3 doc/tickets/build.py` after editing `register.py`.\n")
md.append(f"**{len(R.JOURNEYS)} journeys - {len(set(o[0] for o in R.OBS))} tickets - "
          f"{n_obs} observations - {n_closed} closed - {n_open} open**\n")
md.append("| Bucket | Count | Who closes it |")
md.append("| --- | ---: | --- |")
for k in ORDER:
    md.append(f"| {STATE[k][0]} | {tot[k]} | {OWNER[k]} |")
md.append("")
md.append("## Waiting on an abapGit Pull\n")
md.append("A code fix is not live until the object is pulled AND activated. In the pull "
          "dialog every *Overwrite local object* row arrives **unticked**.\n")
md.append("| Object | What changed | Ticket |")
md.append("| --- | --- | --- |")
for o, w, t in R.PENDING_ACTIVATION:
    md.append(f"| `{o}` | {w} | {t} |")
md.append("")
for code, name, dept, tk, jst, who, note in R.JOURNEYS:
    rows = by_j.get(code, [])
    c = counts([(0,0,0,r[3],0) for r in rows])
    cl = c["fw"] + c["done_dev"]
    md.append(f"## {code} - {name}\n")
    md.append(f"{dept} · {tk} · {jst} · {who} · blocks {R.BLOCKS.get(tk,'-')} · "
              f"{len(R.ATTACHMENTS.get(tk,[]))} screenshots on the ticket · "
              f"**{len(rows)} observations, {cl} closed, {len(rows)-cl} open**\n")
    md.append(note + "\n")
    md.append("| # | Observation | Status | What closes it |")
    md.append("| ---: | --- | --- | --- |")
    for t, seq, text, st, act in rows:
        md.append(f"| {seq} | {text} | {STATE[st][0]} | {act} |")
    md.append("")
open(os.path.join(HERE, "SNAPSHOT.md"), "w").write("\n".join(md) + "\n")

# ---------------------------------------------------------------- html
e = html.escape
def chip(st):
    lab, cls = STATE[st]
    return f'<span class="chip {cls}">{lab}</span>'

parts = []
A = parts.append
A('<title>CJS Journey Board</title>')
A('<link rel="stylesheet" href="https://fonts.googleapis.com/css2?'
  'family=Archivo:wght@500;600;700&family=IBM+Plex+Mono:wght@400;500&'
  'family=IBM+Plex+Sans:wght@400;500;600&display=swap">')
A("<style>%CSS%</style>")

A('<header class="top">')
A('<div class="wrap">')
A('<p class="eyebrow">RAK Customer Journey Studio · SIT round</p>')
A('<h1>Journey board</h1>')
A(f'<p class="lede">Every observation the testers raised across {len(set(o[0] for o in R.OBS))} '
  f'Jira tickets, said plainly, with who closes it and how. As at {AS_AT}.</p>')
A('<dl class="figs">')
for lab, val in (("Journeys", len(R.JOURNEYS)),
                 ("Tickets", len(set(o[0] for o in R.OBS))),
                 ("Observations", n_obs),
                 ("Closed", n_closed),
                 ("Open", n_open)):
    A(f'<div><dt>{lab}</dt><dd>{val}</dd></div>')
A('</dl>')
A('<div class="bar" role="img" aria-label="Where the work sits">')
for k in ORDER:
    pct = tot[k] * 100.0 / n_obs
    A(f'<span class="seg {STATE[k][1]}" style="width:{pct:.2f}%" title="{STATE[k][0]}: {tot[k]}"></span>')
A('</div>')
A('<ul class="legend">')
for k in ORDER:
    A(f'<li><i class="sw {STATE[k][1]}"></i>{STATE[k][0]} <b>{tot[k]}</b></li>')
A('</ul>')
A('</div></header>')

A('<div class="wrap">')
A('<section class="callout">')
A('<h2>Who pushes what</h2>')
A('<p>Three kinds of change came out of this round, and they do not travel the same way. '
  '<strong>Framework</strong> is the shared engine - one change, every journey feels it - so we make it, '
  'push it, and the team pulls it. <strong>Handler class</strong> is the journey\'s own ABAP and belongs to '
  'the developer who owns that service; where a fix is already written it is sitting in git for them to read, '
  'take or reject - nobody pulls it on their behalf. <strong>Config</strong> never leaves the Studio.</p>')
A('<h3>Framework - made and pushed by us</h3>')
A('<p class="sub">Pull these and activate them. In the abapGit pull dialog every '
  '<em>Overwrite local object</em> row arrives unticked and the ticks reset each time the dialog '
  'opens, so an existing class is skipped unless you tick it by hand. Take '
  '<code>ZCL_RAK_JOURNEY_LOGIC</code> and <code>ZCL_RAK_JOURNEY_RENDER</code> first; they break widest.</p>')
A('<table class="pend"><thead><tr><th>Object</th><th>What changed</th></tr></thead><tbody>')
for o, w in R.FRAMEWORK:
    A(f'<tr><td><code>{e(o)}</code></td><td>{e(w)}</td></tr>')
A('</tbody></table>')
hrows = [(o[2], o[0], o[3], o[5]) for o in R.OBS if o[4] == "handler"]
A('<h3>Handler class - written, but the developer\'s call</h3>')
A('<p class="sub">These are in git too, on the same branch. They are one journey\'s own class, so read '
  'the diff before taking it - and if you would rather write it yourself, the observation column says '
  'exactly what the defect was.</p>')
A('<table class="pend"><thead><tr><th>Journey</th><th>Class</th><th>What it fixes</th></tr></thead><tbody>')
for j, tk, what, act in hrows:
    A(f'<tr><td class="num">{j} · {tk}</td><td><code>{e(R.HANDLER.get(j,"-").split(" (")[0])}</code></td>'
      f'<td>{e(what)}</td></tr>')
A('</tbody></table>')
A('</section>')

A('<div class="filters" id="filters">')
A('<input type="search" id="q" placeholder="Search observations, journeys, field names" aria-label="Search">')
A('<div class="chips" id="stchips">')
for k in ORDER:
    A(f'<button class="fchip {STATE[k][1]}" data-st="{k}" aria-pressed="false">{STATE[k][0]}'
      f'<b>{tot[k]}</b></button>')
A('</div>')
A('<div class="chips">')
for d in ("DOK","EPDA"):
    A(f'<button class="fchip dept" data-dept="{d}" aria-pressed="false">{d}</button>')
A('<button class="fchip clear" id="clear">Clear</button>')
A('</div>')
A('</div>')
A('<p class="shown" id="shown"></p>')

for code, name, dept, tk, jst, who, note in R.JOURNEYS:
    rows = by_j.get(code, [])
    c = counts([(0,0,0,r[3],0) for r in rows])
    cl = c["fw"] + c["done_dev"]
    A(f'<section class="jny" data-dept="{dept}" data-code="{code}">')
    A('<div class="jhead">')
    A(f'<h2><span class="jcode">{code}</span> {e(name)}</h2>')
    A(f'<p class="meta"><span class="dept {dept}">{dept}</span>'
      f'<span>{tk}</span><span>{jst}</span><span>{e(who)}</span>'
      f'<span title="blocks the delivery ticket">&#8594; {R.BLOCKS.get(tk,"-")}</span></p>')
    A('</div>')
    A('<div class="prog"><div class="pbar">')
    for k in ORDER:
        if c[k]:
            A(f'<span class="seg {STATE[k][1]}" style="width:{c[k]*100.0/len(rows):.2f}%"></span>')
    A('</div>')
    A(f'<p class="pnum"><b>{cl}</b> of {len(rows)} closed</p></div>')
    A('<p class="split">')
    A(f'<span class="k">Handler</span> <code>{e(R.HANDLER.get(code,"-"))}</code>')
    A(f'<span class="k">Config</span> {c["config"]} point{"s" if c["config"]!=1 else ""}'
      f' &#183; <span class="k">Not CJS</span> {c["backend"]+c["portal"]}'
      f' &#183; <span class="k">Needs a look</span> {c["check"]}</p>')
    A(f'<p class="note">{e(note)}</p>')
    shots = R.ATTACHMENTS.get(tk, [])
    if shots:
        A(f'<p class="shots" title="{e(chr(10).join(shots))}">{len(shots)} screenshot'
          f'{"s" if len(shots)>1 else ""} attached to {tk} &#183; not reachable from here, '
          f'paste one in if a line below needs it</p>')
    A('<table class="obs"><thead><tr><th class="n">#</th><th>Observation</th>'
      '<th>Status</th><th>What closes it</th></tr></thead><tbody>')
    for t, seq, text, st, act in rows:
        hay = e((text + " " + act + " " + code + " " + name + " " + t).lower())
        A(f'<tr data-st="{st}" data-hay="{hay}">'
          f'<td class="n">{seq}</td><td class="what">{e(text)}</td>'
          f'<td>{chip(st)}</td><td class="how">{e(act)}</td></tr>')
    A('</tbody></table>')
    A('</section>')

A('<footer><p>Source: Jira exports of 5 September 2026, Issues and Comments sheets, five assignees. '
  'The register lives in the repository at <code>doc/tickets/register.py</code>; this page is '
  'generated from it by <code>doc/tickets/build.py</code>, so the next round is an edit and a re-run.</p>'
  '<p>Nothing on this page has been compiled or activated - there is no SAP connection from where it '
  'was written. Code marked fixed is fixed in git.</p></footer>')
A('</div>')
A("<script>%JS%</script>")

CSS = r"""
:root{
  --ground:#EEF2F2; --surface:#FFFFFF; --surface2:#F6F9F9; --raise:#FBFDFD;
  --ink:#0F1A1B; --ink2:#4B5C5E; --ink3:#6E8082; --rule:#D5DEDE; --rule2:#E6ECEC;
  --accent:#0B6E68;
  --fw:#0B6E68; --hd:#1C6B43; --done:#4F7A62; --cfg:#9E5A08; --be:#5B4B8A;
  --pt:#4A5A72; --chk:#9C3A52;
  --fwb:#DDEDEC; --hdb:#E4F0E9; --doneb:#EBF1EE; --cfgb:#F7EBDC; --beb:#EBE7F4;
  --ptb:#E7EBF1; --chkb:#F7E5E9;
}
@media (prefers-color-scheme:dark){
  :root:not([data-theme="light"]){
    --ground:#0D1314; --surface:#141C1D; --surface2:#182122; --raise:#1C2627;
    --ink:#E3ECEB; --ink2:#A2B3B3; --ink3:#7E9091; --rule:#263234; --rule2:#1F2A2B;
    --accent:#4FBDB2;
    --fw:#4FBDB2; --hd:#66C08D; --done:#8FB3A0; --cfg:#E0A54E; --be:#A99AE0;
    --pt:#9DAFC9; --chk:#E88CA0;
    --fwb:#123030; --hdb:#162A20; --doneb:#1A2621; --cfgb:#2C2113; --beb:#211C33;
    --ptb:#1A222C; --chkb:#2E1A20;
  }
}
:root[data-theme="dark"]{
  --ground:#0D1314; --surface:#141C1D; --surface2:#182122; --raise:#1C2627;
  --ink:#E3ECEB; --ink2:#A2B3B3; --ink3:#7E9091; --rule:#263234; --rule2:#1F2A2B;
  --accent:#4FBDB2;
  --fw:#4FBDB2; --hd:#66C08D; --done:#8FB3A0; --cfg:#E0A54E; --be:#A99AE0;
  --pt:#9DAFC9; --chk:#E88CA0;
  --fwb:#123030; --hdb:#162A20; --doneb:#1A2621; --cfgb:#2C2113; --beb:#211C33;
  --ptb:#1A222C; --chkb:#2E1A20;
}
*{box-sizing:border-box}
body{background:var(--ground);color:var(--ink);
  font-family:"IBM Plex Sans",-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;
  font-size:15px;line-height:1.55;-webkit-font-smoothing:antialiased}
.wrap{max-width:1120px;margin:0 auto;padding:0 24px}
code{font-family:"IBM Plex Mono",ui-monospace,SFMono-Regular,Menlo,monospace;
  font-size:.88em;color:var(--ink);background:var(--surface2);
  padding:1px 5px;border-radius:3px;border:1px solid var(--rule2)}
h1,h2,h3{font-family:Archivo,"IBM Plex Sans",sans-serif;text-wrap:balance;margin:0}
.top{background:var(--surface);border-bottom:1px solid var(--rule);padding:44px 0 28px}
.eyebrow{font-size:11.5px;letter-spacing:.14em;text-transform:uppercase;
  color:var(--accent);font-weight:600;margin:0 0 10px}
h1{font-size:clamp(30px,4.4vw,46px);font-weight:700;letter-spacing:-.02em;line-height:1.04}
.lede{max-width:62ch;color:var(--ink2);margin:12px 0 0;font-size:16px}
.figs{display:flex;flex-wrap:wrap;gap:36px;margin:28px 0 0}
.figs div{display:flex;flex-direction:column-reverse}
.figs dt{font-size:11.5px;letter-spacing:.09em;text-transform:uppercase;color:var(--ink3);margin:2px 0 0}
.figs dd{margin:0;font-family:Archivo,sans-serif;font-weight:600;font-size:28px;
  font-variant-numeric:tabular-nums;line-height:1}
.bar{display:flex;height:10px;border-radius:5px;overflow:hidden;margin:26px 0 12px;
  background:var(--surface2)}
.seg{display:block;height:100%}
.seg.fw{background:var(--fw)} .seg.hd{background:var(--hd)} .seg.done{background:var(--done)} .seg.cfg{background:var(--cfg)}
.seg.be{background:var(--be)} .seg.pt{background:var(--pt)} .seg.chk{background:var(--chk)}
.legend{list-style:none;display:flex;flex-wrap:wrap;gap:6px 22px;margin:0;padding:0;
  font-size:13px;color:var(--ink2)}
.legend li{display:flex;align-items:center;gap:7px}
.legend b{font-variant-numeric:tabular-nums;color:var(--ink)}
.sw{width:9px;height:9px;border-radius:2px;display:inline-block}
.sw.fw{background:var(--fw)} .sw.hd{background:var(--hd)} .sw.done{background:var(--done)} .sw.cfg{background:var(--cfg)}
.sw.be{background:var(--be)} .sw.pt{background:var(--pt)} .sw.chk{background:var(--chk)}
.callout{margin:36px 0 0;padding:22px 24px;background:var(--surface);
  border:1px solid var(--rule);border-left:3px solid var(--accent);border-radius:2px}
.callout h2{font-size:19px;font-weight:600;letter-spacing:-.01em}
.callout p{color:var(--ink2);max-width:74ch;margin:8px 0 16px}
table{border-collapse:collapse;width:100%}
.pend{font-size:13.5px}
.pend th{text-align:left;font-weight:600;font-size:11px;letter-spacing:.09em;
  text-transform:uppercase;color:var(--ink3);padding:0 12px 6px 0;border-bottom:1px solid var(--rule)}
.pend td{padding:7px 12px 7px 0;border-bottom:1px solid var(--rule2);vertical-align:top;color:var(--ink2)}
.pend td.num{font-family:"IBM Plex Mono",monospace;font-size:12px;white-space:nowrap;color:var(--ink3)}
.filters{position:sticky;top:0;z-index:5;background:var(--ground);
  padding:18px 0 12px;margin:36px 0 0;border-bottom:1px solid var(--rule);
  display:flex;flex-wrap:wrap;gap:10px 16px;align-items:center}
#q{flex:1 1 280px;min-width:220px;padding:9px 12px;border:1px solid var(--rule);
  border-radius:3px;background:var(--surface);color:var(--ink);font:inherit;font-size:14px}
#q:focus-visible,.fchip:focus-visible{outline:2px solid var(--accent);outline-offset:2px}
.chips{display:flex;flex-wrap:wrap;gap:7px}
.fchip{font:inherit;font-size:12.5px;padding:6px 11px;border-radius:99px;cursor:pointer;
  background:var(--surface);color:var(--ink2);border:1px solid var(--rule);
  display:inline-flex;align-items:center;gap:7px}
.fchip b{font-variant-numeric:tabular-nums;color:var(--ink3);font-weight:500}
.fchip[aria-pressed="true"]{background:var(--ink);color:var(--ground);border-color:var(--ink)}
.fchip[aria-pressed="true"] b{color:var(--ground);opacity:.7}
.shown{font-size:12.5px;color:var(--ink3);margin:12px 0 0;height:1em}
.jny{margin:34px 0 0;background:var(--surface);border:1px solid var(--rule);border-radius:2px;
  padding:22px 24px 6px}
.jhead{display:flex;flex-wrap:wrap;gap:6px 18px;align-items:baseline;justify-content:space-between}
.jny h2{font-size:21px;font-weight:600;letter-spacing:-.01em}
.jcode{font-family:"IBM Plex Mono",monospace;font-size:16px;font-weight:500;
  color:var(--accent);margin-right:8px}
.meta{display:flex;gap:14px;margin:0;font-size:12.5px;color:var(--ink3);
  font-family:"IBM Plex Mono",monospace}
.dept{font-weight:500;color:var(--ink2)}
.prog{display:flex;align-items:center;gap:14px;margin:16px 0 0}
.pbar{display:flex;height:6px;flex:1;border-radius:3px;overflow:hidden;background:var(--surface2)}
.pnum{margin:0;font-size:12.5px;color:var(--ink3);white-space:nowrap;font-variant-numeric:tabular-nums}
.pnum b{color:var(--ink)}
.callout h3{font-size:14px;font-weight:600;margin:22px 0 0;letter-spacing:.01em}
.callout .sub{font-size:13.5px;margin:6px 0 12px}
.split{margin:14px 0 0;font-size:12.5px;color:var(--ink3);display:flex;flex-wrap:wrap;gap:6px 14px;align-items:baseline}
.split .k{font-size:10.5px;letter-spacing:.1em;text-transform:uppercase;color:var(--ink3);font-weight:600}
.split code{font-size:11.5px}
.shots{margin:8px 0 0;font-size:12.5px;color:var(--ink3)}
.note{color:var(--ink2);max-width:78ch;margin:14px 0 4px;font-size:14.5px}
.obs{font-size:14px;margin:10px 0 0}
.obs th{text-align:left;font-weight:600;font-size:11px;letter-spacing:.09em;text-transform:uppercase;
  color:var(--ink3);padding:10px 14px 7px 0;border-bottom:1px solid var(--rule)}
.obs td{padding:11px 14px 11px 0;border-bottom:1px solid var(--rule2);vertical-align:top}
.obs tr:last-child td{border-bottom:0}
.obs .n{width:26px;color:var(--ink3);font-variant-numeric:tabular-nums;font-size:12.5px;padding-top:12px}
.obs .what{width:34%;color:var(--ink)}
.obs .how{color:var(--ink2);font-size:13.5px}
.obs td:nth-child(3){width:172px}
.chip{display:inline-block;font-size:11.5px;font-weight:500;padding:3px 9px;border-radius:99px;
  white-space:nowrap;line-height:1.4}
.chip.fw{background:var(--fwb);color:var(--fw)}
.chip.hd{background:var(--hdb);color:var(--hd)}
.chip.done{background:var(--doneb);color:var(--done)}
.chip.cfg{background:var(--cfgb);color:var(--cfg)}
.chip.be{background:var(--beb);color:var(--be)}
.chip.pt{background:var(--ptb);color:var(--pt)}
.chip.chk{background:var(--chkb);color:var(--chk)}
.jny.hide,.obs tr.hide{display:none}
footer{margin:48px 0 60px;padding:20px 0 0;border-top:1px solid var(--rule);
  font-size:13px;color:var(--ink3);max-width:80ch}
footer p{margin:0 0 8px}
@media (max-width:720px){
  .obs td:nth-child(3){width:auto}
  .obs .what{width:auto}
  .obs,.pend{display:block;overflow-x:auto}
  .figs{gap:24px}
}
@media (prefers-reduced-motion:reduce){*{transition:none!important;animation:none!important}}
"""

JS = r"""
(function(){
  var q=document.getElementById('q'), shown=document.getElementById('shown');
  var st=new Set(), dept=new Set();
  function chips(){return Array.prototype.slice.call(document.querySelectorAll('.fchip[data-st],.fchip[data-dept]'));}
  function apply(){
    var term=(q.value||'').trim().toLowerCase(), vis=0, tot=0;
    document.querySelectorAll('.jny').forEach(function(j){
      var any=false;
      if(dept.size && !dept.has(j.dataset.dept)){ j.classList.add('hide');
        j.querySelectorAll('tbody tr').forEach(function(r){tot++;}); return; }
      j.querySelectorAll('tbody tr').forEach(function(r){
        tot++;
        var ok=(!st.size||st.has(r.dataset.st)) && (!term||r.dataset.hay.indexOf(term)>-1);
        r.classList.toggle('hide',!ok); if(ok){any=true;vis++;}
      });
      j.classList.toggle('hide',!any);
    });
    shown.textContent = (st.size||dept.size||term) ? ('Showing '+vis+' of '+tot+' observations') : '';
  }
  chips().forEach(function(b){
    b.addEventListener('click',function(){
      var on=b.getAttribute('aria-pressed')==='true';
      b.setAttribute('aria-pressed',on?'false':'true');
      var set=b.dataset.st?st:dept, key=b.dataset.st||b.dataset.dept;
      if(on){set.delete(key);}else{set.add(key);}
      apply();
    });
  });
  q.addEventListener('input',apply);
  document.getElementById('clear').addEventListener('click',function(){
    st.clear();dept.clear();q.value='';
    chips().forEach(function(b){b.setAttribute('aria-pressed','false');});
    apply();
  });
})();
"""

out = "\n".join(p for p in parts if p)
out = out.replace("%CSS%", CSS).replace("%JS%", JS)
open(os.path.join(HERE, "snapshot.html"), "w").write(out)
json.dump({"as_at": AS_AT,
           "journeys": [dict(zip(("code","name","dept","ticket","jira","assignee","note"), j)) for j in R.JOURNEYS],
           "observations": [dict(zip(("ticket","seq","journey","observation","state","action"), o)) for o in R.OBS],
           "pending_activation": [dict(zip(("object","change","ticket"), p)) for p in R.PENDING_ACTIVATION],
           "blocks": R.BLOCKS, "attachments": R.ATTACHMENTS},
          open(os.path.join(HERE, "register.json"), "w"), indent=1, ensure_ascii=False)
print("SNAPSHOT.md, snapshot.html, register.json written")

# ---------------------------------------------------------------- xlsx
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

if Workbook:
    HEAD = PatternFill("solid", fgColor="0B3B39")
    HF   = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    BODY = Font(size=10, name="Calibri")
    MONO = Font(size=9, name="Consolas")
    WRAP = Alignment(wrap_text=True, vertical="top")
    TOP  = Alignment(vertical="top")
    THIN = Border(bottom=Side(style="thin", color="D9D9D9"))
    TINT = {"fw":"DCEDEB","handler":"E2F0E6","done_dev":"EFF3F1","config":"FBEEDD",
            "backend":"ECE8F4","portal":"E8ECF2","check":"F9E7EA"}

    def head(ws, row, cols):
        for i, (t, w) in enumerate(cols, 1):
            c = ws.cell(row=row, column=i, value=t)
            c.fill, c.font, c.alignment = HEAD, HF, Alignment(wrap_text=True, vertical="center")
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[row].height = 30

    wb = Workbook()

    # ---- sheet 1: one row per journey, split by who does it
    ws = wb.active; ws.title = "By journey"
    ws["A1"] = "CJS journeys - who does what, %s" % AS_AT
    ws["A1"].font = Font(bold=True, size=13, name="Calibri")
    ws["A2"] = ("Framework changes are made and pushed by the CJS team - pull and activate them. "
                "Handler and config work stays with the journey's own developer.")
    ws["A2"].font = Font(size=10, italic=True, color="5A6A6C", name="Calibri")
    cols = [("Journey",9),("Service",30),("Dept",7),("Ticket",13),("Blocks",13),
            ("Assignee",14),("Obs",6),("Closed",7),("Open",6),
            ("FRAMEWORK - done by us, pull it",44),
            ("HANDLER CLASS - developer",46),
            ("CONFIG in the Studio - developer",52),
            ("BACKEND / PORTAL - other teams",34),
            ("NEEDS A LOOK",30)]
    head(ws, 4, cols)
    ws.freeze_panes = "A5"
    r = 5
    for code, name, dept, tk, jst, who, note in R.JOURNEYS:
        rows = by_j.get(code, [])
        c = counts([(0,0,0,x[3],0) for x in rows])
        cl = c["fw"] + c["done_dev"]
        def lines(state):
            return "\n".join("%d. %s" % (x[1], x[2]) for x in rows if x[3] == state)
        fwtxt = lines("fw") or "-"
        hdtxt = R.HANDLER.get(code, "-")
        w = lines("handler")
        if w:
            hdtxt += "\nFIX WRITTEN, IN GIT - YOURS TO TAKE:\n" + w
        cfg = lines("config") or "-"
        ext = "\n".join(x for x in (lines("backend"), lines("portal")) if x) or "-"
        chk = lines("check") or "-"
        vals = [code, name, dept, tk, R.BLOCKS.get(tk, ""), who,
                len(rows), cl, len(rows) - cl, fwtxt, hdtxt, cfg, ext, chk]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font = MONO if i == 11 else BODY
            cell.alignment = WRAP if i >= 10 or i == 2 else TOP
            cell.border = THIN
        if fwtxt != "-":
            ws.cell(row=r, column=10).fill = PatternFill("solid", fgColor=TINT["fw"])
        if w:
            ws.cell(row=r, column=11).fill = PatternFill("solid", fgColor=TINT["handler"])
        ws.cell(row=r, column=12).fill = PatternFill("solid", fgColor=TINT["config"])
        r += 1

    # ---- sheet 2: every observation
    ws2 = wb.create_sheet("All observations")
    cols2 = [("Journey",9),("Service",26),("Ticket",13),("#",5),("Observation",56),
             ("Who does it",26),("Bucket",22),("What closes it",76)]
    head(ws2, 1, cols2); ws2.freeze_panes = "A2"
    r = 2
    jname = {j[0]: j[1] for j in R.JOURNEYS}
    for tk, seq, j, text, st, act in R.OBS:
        vals = [j, jname.get(j, ""), tk, seq, text, OWNER[st], STATE[st][0], act]
        for i, v in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=i, value=v)
            cell.font = BODY; cell.alignment = WRAP if i in (5, 8) else TOP
            cell.border = THIN
        for i in (6, 7):
            ws2.cell(row=r, column=i).fill = PatternFill("solid", fgColor=TINT[st])
        r += 1
    ws2.auto_filter.ref = "A1:H%d" % (r - 1)

    # ---- sheet 3: the framework changes, on their own
    ws3 = wb.create_sheet("Framework changes")
    ws3["A1"] = "Framework - made and pushed by the CJS team"
    ws3["A1"].font = Font(bold=True, size=13, name="Calibri")
    ws3["A2"] = ("Pull and activate. Every 'Overwrite local object' row in the abapGit pull dialog "
                 "arrives UNTICKED and the ticks reset each time it opens - tick them by hand. "
                 "ZCL_RAK_JOURNEY_LOGIC and ZCL_RAK_JOURNEY_RENDER first.")
    ws3["A2"].font = Font(size=10, italic=True, color="5A6A6C", name="Calibri")
    ws3["A2"].alignment = WRAP
    head(ws3, 4, [("Object",34),("What changed",100),("Applies to",24)])
    ws3.freeze_panes = "A5"
    r = 5
    for o, w in R.FRAMEWORK:
        scope = "run it per journey" if o.startswith("ZRAK_") else "every journey"
        for i, v in enumerate([o, w, scope], 1):
            cell = ws3.cell(row=r, column=i, value=v)
            cell.font = MONO if i == 1 else BODY
            cell.alignment = WRAP; cell.border = THIN
        r += 1

    wb.save(os.path.join(HERE, "CJS_journey_ownership.xlsx"))
    print("CJS_journey_ownership.xlsx written")
